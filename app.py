from flask import Flask
import json
import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from random import randint
import pandas as pd
from flask_cors import CORS
import datetime
from datetime import datetime
import win32com.client as win32
from win32com.client import Dispatch

app=Flask(__name__)
CORS(app)
# f = open("quotes.json", "r")
#open file quotes and get random quote.
def random_qoute():
    with open('quotes.json') as rq:
        data = json.load(rq)
        qoute = data["quotes"]
        quoteLen=len(data["quotes"])
       
        random_index = randint(0, len(qoute)-1)
        randomQuote=data["quotes"][random_index]
        quoteId=randomQuote["id"]

         #open excel file and update count value
        # wb = xlrd.open_workbook('Quotes Report.xlsx')
        wb2=openpyxl.load_workbook('Quotes Report_Res2.xlsx')
        sheets = wb2.sheetnames
        sheet=wb2[sheets[0]]
        for i in range(len(qoute)):
            if quoteId==sheet.cell(row=i+2,column=1).value:
                sheet.cell(row=i+2,column=2).value+=1 
        # print(wb2) 
        now=datetime.now()
        print(now)
        # convert datetime obj to string
        # current_datetime = str(now)
        # fileName=current_datetime+'.xlsx'
        wb2.save('Quotes Report_Res2.xlsx')   #it saves only one if not change name after update 
        return randomQuote

#find quote if i have the id;
def search_quote(id):
    with open('quotes.json') as rq:
        # get all content of a file
        data = json.load(rq)
        quoteLen=len(data["quotes"])
        for i in range(quoteLen):
            if(id==data["quotes"][i]["id"]):
                qouteVal=data["quotes"][i]["quote"]
                return qouteVal
        
#open file authors and get author of the quote.
def quote_details():
    with open('authors.json') as sa:
        data = json.load(sa)
        authLen=len(data["authors"])
        authors = data["authors"]
        quoteID=random_qoute()["id"] 
        print(quoteID)
        result={
            "quoteId": 0,
            "quote": "",
            "author": ""
        }
        for i in range(authLen):
            if quoteID in authors[i]["quoteIds"]:
                result["quoteId"]=quoteID
                result["quote"]=search_quote(quoteID)
                result["author"]=authors[i]["author"]
                return result

# create Excel File and add column , run once       
# def create_Excel():
#     wb = load_workbook('Quotes Report_new.xlsx') 
#     sheets = wb.sheetnames
    
#     sheet=wb[sheets[0]]
#     sheet.cell(row=1,column=1).value="id"
#     sheet.cell(row=1,column=2).value="Count"
   
#     with open('quotes.json') as rq:
#         data = json.load(rq)
#         qoute = data["quotes"]
#         my_list =[]
#         for i in range(len(qoute)):
#             id=data["quotes"][i]["id"]
#             my_list.append({"id":id,"count":0})
#     for i in range(len(my_list)):                                          
#          sheet.cell(row=i+2,column=1).value=my_list[i]["id"]                                      
#          sheet.cell(row=i+2,column=2).value=0      
#     wb.save('Quotes Report_Res.xlsx')
       
    
    

@app.route('/quote/random',methods=['GET'])
def index():
    # create_Excel()
    return quote_details()

if __name__=="__main__":
    app.run(debug=True)

